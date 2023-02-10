from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['C1'] = 'Gender'

treeData = [["Type", "Leaf Color", "Height"], ["Maple", "Red", 549], ["Oak", "Green", 783], ["Pine", "Green", 1204]]
print(type(treeData))

quit()


# Rows can also be appended
x = 0
while x < 3:
    name = input("Name: ")
    age = input("Age: ")
    ws.append([name, age])
    x = x + 1

x = 0
while x < 3:
    gender = input("Gender: ")
    ws.append([gender],)
    x = x + 1

# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
