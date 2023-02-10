from bs4 import BeautifulSoup
import requests
import re
import openpyxl as xl

wb = xl.Workbook()
ws = wb.active
ws['A1'] = 'Title'
ws['B1'] = 'Price'
ws['C1'] = 'Ratings'
URL = input("Enter URL: ")
if len(URL) < 1:
    URL = 'https://www.amazon.com/s?k=power+supply&qid=1675700171&ref=sr_pg_1'

'       Find your user-agent on this website -> https://www.whatismybrowser.com/detect/what-is-my-user-agent/       '
HEADERS = ({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/109.0',
            'Accept-Language': 'en-US, en;q=0.5'})  # add your user agent
webpage = requests.get(URL, headers=HEADERS)
soup = BeautifulSoup(webpage.content, 'html.parser')

titleTag = soup.find_all('span', attrs={'class': 'a-size-medium a-color-base a-text-normal'})
ratingsTag = soup.find_all('span', attrs={'class': 'a-icon-alt'})
priceTag = soup.find_all('span', attrs={'class': 'a-offscreen'})


titles = list()
#                                                       'This part will fetch all the titles'
i = 0
for title in titleTag:
    x = str(title)
    x = re.findall('>(.+)<', x)
    # print('Title: ', x)
    titles.append(x)
    i = i + 1
    if i == 20:
        break
prices = list()
#                                                       'This part will fetch all the prices'
i = 0
for price in priceTag:
    x = str(price)
    x = re.findall('>(.+)<', x)
    # print('Price: ', x)
    prices.append(x)
    i = i + 1
    if i == 20:
        break

ratings = list()
#                                                       'This part will fetch all the ratings'
i = 0
for rating in ratingsTag:
    x = str(rating)
    x = re.findall('>(.+)<', x)
    # print('Ratings: ', x)
    ratings.append(x)
    i = i + 1
    if i == 20:
        break
# print(titles[0])
# print(prices)
# print(ratings)
i = 0
while True:
    x = str(titles[i])
    y = str(prices[i])
    z = str(ratings[i])
    y = y[2:-2]
    x = x[2:-2]
    z = z[2:-2]
    # print(x)
    # print(y)
    # print(z)
    # break
    ws.append([x, y, z])
    # ws.append([1,2,3])
    i = i + 1
    if i == 20:
        break

wb.save('sample.xlsx')
